import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
import os

# Load the employee data
input_file = "Planilla para generar comprobantes de pago.xlsx"
data = pd.read_excel(input_file)

# Output folder
output_folder = "Recibos"
os.makedirs(output_folder, exist_ok=True)

# Define function to generate a receipt for each employee
def generate_receipt(employee):
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Boleta de Pago"

    # Styles
    title_font = Font(size=14, bold=True)
    bold_font = Font(bold=True)
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal="center")

    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "BOLETA DE PAGO QUINCENA"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_alignment

    ws.merge_cells("A2:E2")
    ws["A2"] = "Correspondiente del 16 al 29 de Febrero 2024"
    ws["A2"].font = bold_font
    ws["A2"].alignment = center_alignment

    # Employee details
    ws["A4"] = "Nombre del Empleado"
    ws["B4"] = employee["Nombre"]
    ws["A5"] = "IGSS:"
    ws["B5"] = employee["IGSS"]
    ws["A6"] = "No. De Antigüedad"
    ws["B6"] = employee["Antigüedad"]
    ws["A7"] = "Días Laborados"
    ws["B7"] = employee["Dias_Laborados"]

    # Payment breakdown
    ws["A9"] = "Desglose de Pago"
    ws["A10"] = "Ingresos"
    ws["C10"] = "Descuentos"

    ws["A11"] = "Sueldo Ordinario"
    ws["B11"] = employee["Sueldo_Ordinario"]
    ws["C11"] = "Cuota Laboral IGSS"
    ws["D11"] = employee["Cuota_Laboral_IGSS"]

    ws["A12"] = "Bonificación de ley"
    ws["B12"] = employee["Bonificacion_Ley"]
    ws["C12"] = "Otros Descuentos"
    ws["D12"] = employee["Otros_Descuentos"]

    ws["A13"] = "Bonificación incentivo"
    ws["B13"] = employee["Bonificacion_Incentivo"]
    ws["C13"] = "ISR empleado"
    ws["D13"] = employee["ISR"]

    ws["A14"] = "Total Ingresos"
    ws["B14"] = employee["Total_Ingresos"]
    ws["C14"] = "Total Egresos"
    ws["D14"] = employee["Total_Egresos"]

    # Net Payment
    ws["A16"] = "Líquido a Recibir"
    ws["B16"] = employee["Liquido_Recibir"]

    # Payment Method
    ws["A18"] = "Forma de Pago"
    ws["B18"] = "Banco"
    ws["C18"] = employee["Banco"]

    # Signature line
    ws["A20"] = "Recibí Conforme"
    ws["B20"] = employee["Cuenta"]
    ws["C20"] = employee["Fecha"]

    ws["A21"] = "Firma"
    ws["B21"] = "DPI"
    ws["C21"] = employee["DPI"]

    # Apply borders
    for row in ws.iter_rows(min_row=4, max_row=21, min_col=1, max_col=5):
        for cell in row:
            cell.border = border

    # Save the receipt
    filename = f"{output_folder}/{employee['Nombre'].replace(' ', '_')}_Boleta_Pago.xlsx"
    wb.save(filename)

# Iterate through each employee and generate receipts
for _, row in data.iterrows():
    generate_receipt(row)